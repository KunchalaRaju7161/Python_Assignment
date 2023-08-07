import openpyxl
from API_Assignment.Test_Data.Config import ConfigItems


def read_excel_file(tasknumber):
    workbook = openpyxl.load_workbook(ConfigItems.EXCEL_PATH)
    worksheet = workbook[ConfigItems.SHEET_NAME]

    dict_from_excel = {}
    for row in worksheet.iter_rows(values_only=True):
        dict_from_excel[row[1]] = row[tasknumber]

    workbook.close()
    return dict_from_excel


def write_to_excel(data):
    workbook = openpyxl.load_workbook(ConfigItems.EXCEL_PATH)
    worksheet = workbook[ConfigItems.SHEET_NAME]

    # Assuming data is a list of values to be written in the row
    worksheet.append(data)

    workbook.save(ConfigItems.EXCEL_PATH)
    workbook.close()
