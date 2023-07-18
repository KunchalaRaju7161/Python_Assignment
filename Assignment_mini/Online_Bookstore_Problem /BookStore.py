# import openpyxl module
import openpyxl

# Give the file location path
path = "/Users/rkunchala/pythonAssignments/Python_Assignments/Assignment_mini/Resource/OnlineBookstore.xlsx"

# To open the workbook
#  object is created
obj = openpyxl.load_workbook(path)
sheet_obj = obj.active
max_col = sheet_obj.max_column

# create empty books array
# By use for loop get the data
books = []
for i in range(2, max_col + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    books.append(cell_obj.value)

# create empty prices array
# By use for loop get the data
prices = []
for i in range(2, max_col + 1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    prices.append(cell_obj.value)

# Threshold for filtering books
threshold = 50


# Filtering books based on the price threshold
def filter_books(books, prices, threshold):
    filtered_books = []
    for book, price in zip(books, prices):
        if price > threshold:
            filtered_books.append(book)
    return filtered_books


# Calculating the total revenue from the sales of the filtered books
def calculate_total_revenue(prices, threshold):
    total_revenue = sum(price for price in prices if price > threshold)
    return total_revenue


filter_book = filter_books(books, prices, threshold)
revenue = calculate_total_revenue(prices, threshold)

# Creating a list of book-price tuples
book_price_tuple = list(zip(books, prices))

# print the data
print("Filtered books : ", filter_book)
print("Total revenue : ", revenue)
print("Book prices tuples : ", book_price_tuple)
