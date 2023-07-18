# import openpyxl module
import openpyxl

# Give the location of the file
path = "/Users/rkunchala/pythonAssignments/Python_Assignments/Assignment_mini/Resource/StudentGradesAnalysis.xlsx"

# To open the workbook
#  object is created
wb_obj = openpyxl.load_workbook(path)


# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column

# create empty Alice array
# Loop will get the all grades of Alice
# Find Alice highest average grade
Array_Alice = []
for i in range(2, max_col + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    Array_Alice.append(cell_obj.value)

Alice_Average = (Array_Alice[0] + Array_Alice[1] + Array_Alice[2]) / 3
print("Student: ", sheet_obj.cell(row=1, column=1).value)
print("Average Grade: ", Alice_Average)

# create empty Bob array
# Loop will get all grades of Bob
# Find Bob highest average grade
Array_Bob = []
for i in range(2, max_col + 1):
    cell_obj1 = sheet_obj.cell(row=2, column=i)
    Array_Bob.append(cell_obj1.value)

Bob_Average = (Array_Bob[0] + Array_Bob[1] + Array_Bob[2]) / 3
print("Student: ", sheet_obj.cell(row=2, column=1).value)
print("Average Grade: ", Bob_Average)

#  create empty Alice array
# Loop will get all grades of Charlie
# Find Charlie highest average grade
Array_Charlie = []
for i in range(2, max_col + 1):
    cell_obj2 = sheet_obj.cell(row=3, column=i)
    Array_Charlie.append(cell_obj2.value)

Charlie_Average = (Array_Charlie[0] + Array_Charlie[1] + Array_Charlie[2]) / 3
print("Student : ", sheet_obj.cell(row=3, column=1).value)
print("Average Grade : ", Charlie_Average)

#  create empty David array
# Loop will get all grades of David
# Find David highest average grade
Array_David = []
for i in range(2, max_col + 1):
    cell_obj3 = sheet_obj.cell(row=4, column=i)
    Array_David.append(cell_obj3.value)

David_Average = (Array_David[0] + Array_David[1] + Array_David[2]) / 3
print("Student : ", sheet_obj.cell(row=4, column=1).value)
print("Average Grade : ", David_Average)

# Find the student with the highest average grade
# print the Result
print("Student with the highest average grade: ")
highest_average_grade = max(Alice_Average, Bob_Average, Charlie_Average, David_Average)

if Alice_Average == highest_average_grade:
    print("Student: ", sheet_obj.cell(row=1, column=1).value)
    print("Average Grade: ", Alice_Average)
elif Bob_Average == highest_average_grade:
    print("Student: ", sheet_obj.cell(row=2, column=1).value)
    print("Average Grade: ", Bob_Average)
elif Charlie_Average == highest_average_grade:
    print("Student : ", sheet_obj.cell(row=3, column=1).value)
    print("Average Grade : ", Charlie_Average)
elif David_Average == highest_average_grade:
    print("Student : ", sheet_obj.cell(row=4, column=1).value)
    print("Average Grade : ", David_Average)
