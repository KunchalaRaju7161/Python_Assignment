# create RegisterNewUser class
import openpyxl


class RegisterNewUser():
    # Employee records dictionary to store employee details
    employee_records = {}

    def register_screen(self):
        print("******************** Welcome To User Registration Page **************")
        print("------------------- New User Registration Form -------------")
        RegisterNewUser().new_user_register()

    def new_user_register(self):
        while True:
            print("Employee Registration")
            self.user_id = input("Enter user-id : ")
            # check if user-id already exists
            if self.user_id in self.employee_records:
                print("user -id already exists. Please use a different User -id")
                return
            user_name = input("Enter user name : ")
            password = input("Enter user password : ")
            if not self.validate_password(password):
                print("Invalid Password!")
                print("password should be in the range of 5 to 10 characters")
                print("at lest one uppercase, one lowercase, one digit, and one special character")
                return
            confirm_password = input("Enter user confirm password : ")

            if password == confirm_password:
                print("Registration successful")
                RegisterNewUser().register(user_name,password)
                break
            else:
                print("Registration unsuccessful")
                print("Your provided wrong details")

    def validate_age(self, age):
        return 18 <= age <= 90

    def validate_salary(self, salary):
        return salary > 0

    def validate_phone_number(self, phone_number):
        # check if the phone number contains exactly 10 digits
        if len(phone_number) != 10:
            return False
        # check if all characters in phone number are digits
        if not phone_number.isdigit():
            return False

        return True

    def validate_password(self, password):
        # password should be in the range of 5 to 10 characters
        if 5 <= len(password) <= 10:
            # check for at lest one uppercase, one lowercase, one digit, and one special character
            if any(char.isupper() for char in password) and \
                    any(char.islower() for char in password) and \
                    any(char.isdigit() for char in password) and \
                    any(char in "@#$&*!" for char in password):
                return True
        return False

    def register(self, username: str, passcode: str):
        # Get user information
        print("Kindly provide more info to finish the registration:")
        self.name = input("Name      :>")
        self.email = input("password     :>")


        # Check if the Excel file exists or create a new one
        try:
            workbook = openpyxl.load_workbook('user_credentials.xlsx')
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Select the active worksheet (first sheet by default)
        worksheet = workbook.active

        # Append data to the worksheet
        l = [username, passcode]
        worksheet.append(l)

        # Save the workbook
        workbook.save('user_credentials.xlsx')

        print("Hi", self.name, ", your account has been registered successfully!!!")

    def readUserData(self, username: str, datatype: str):
        try:
            workbook = openpyxl.load_workbook('user_credentials.xlsx')
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=1, values_only=True):
                if username == row[0]:
                    if datatype.upper().__contains__("NAME"):
                        return row[2]
                    elif datatype.upper().__contains__("PASSWORD"):
                        return row[3]
                    # elif datatype.upper().__contains__("PHONE"):
                    #     return row[4]
                    # elif datatype.upper().__contains__("AGE"):
                    #     return row[5]

        except FileNotFoundError:
            print("Database file not found.")
            return None

        print("User not found in the database.")
        return None

