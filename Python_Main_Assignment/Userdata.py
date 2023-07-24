import re

import openpyxl
from openpyxl import Workbook, load_workbook


class Userdata():
    wb = openpyxl.load_workbook('employee_data.xlsx')
    ws = wb.active

    def validate_user_id(self, user_id, employee_data):
        if user_id in employee_data:
            return False
        return True

    def validate_age(self, age):
        if not age.isdigit():
            return False
        age = int(age)
        return 18 <= age <= 90

    def validate_salary(self, salary):
        return salary.isdigit() and int(salary) > 0

    def validate_phone_number(self, phone_number):
        return re.match(r'^\d{10}$', phone_number)

    def validate_password(self, password):
        if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*[@#&%*!])(?=.*\d)[a-zA-Z@#&%*!\d]{5,10}$', password):
            return False
        return True

    def register_employee(self, employee_data):
        user_id = input("Enter User-ID: ")
        if not self.validate_user_id(user_id, employee_data):
            print("User-ID already exists. Try a different one.")
            return

        employee_id = input("Enter Employee Id: ")
        self.name = input("Name      :>")
        self.email = input("Email     :>")
        self.phone = input("PhoneNo   :>")
        self.age = input("Age       :>")
        if not self.validate_age(self.age):
            print("Invalid age. Age must be between 18 and 90.")
            return

        self.company_name = input("Enter Company Name: ")
        self.designation = input("Enter Designation: ")
        self.salary = input("Enter Salary: ")
        if not self.validate_salary(self.salary):
            print("Invalid salary. Salary must be greater than zero.")
            return

        self.address = input("Enter Address: ")
        self.phone_number = input("Enter Phone Number: ")
        if not self.validate_phone_number(self.phone_number):
            print("Invalid phone number. Phone number should be 10 digits only.")
            return

        self.password = input("Enter Password: ")
        if not self.validate_password(self.password):
            print(
                "Invalid password. Password should be 5 to 10 characters with at least one uppercase, one lowercase, one special character, and one digit.")
            return

        employee_data[user_id] = {
            "Employee Id": employee_id,
            "Name": self.name,
            "Age": self.age,
            "Company-Name": self.company_name,
            "Designation": self.designation,
            "Salary": self.salary,
            "Address": self.address,
            "Phone Number": self.phone_number,
            "Password": self.password,
        }
        print("Registration successful.")

    def save_to_excel(self, employee_data):
        # Check if the Excel file exists or create a new one

        try:
            wb = openpyxl.load_workbook('employee_data.xlsx')
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        # Select the active worksheet (first sheet by default)
        ws = wb.active

        # Add headers to the excel sheet
        headers = ["User-ID", "Employee Id", "Name", "Age", "Company-Name", "Designation", "Salary", "Address",
                   "Phone Number", "Password"]
        ws.append(headers)

        # Add employee data to the excel sheet
        for user_id, data in employee_data.items():
            row_data = [user_id] + list(data.values())
            ws.append(row_data)

        wb.save("employee_data.xlsx")

        print("Hi", self.name, ", your account has been registered successfully!!!")

    def readUserData(self, username: str, datatype: str, update_record: str):
        try:
            workbook = openpyxl.load_workbook('employee_data.xlsx')
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if username == row[1]:
                    if datatype.upper().__contains__("User-ID"):
                        return row[2]
                    elif datatype.upper().__contains__("Employe Id"):
                        return row[3]
                    elif datatype.upper().__contains__("Name"):
                        return row[4]
                    elif datatype.upper().__contains__("AGE"):
                        return row[5]
                    elif datatype.upper().__contains__("Company-Name"):
                        return row[6]
                    elif datatype.upper().__contains__("Designation"):
                        return row[7]
                    elif datatype.upper().__contains__("Salary"):
                        return row[8]
                    elif datatype.upper().__contains__("Address"):
                        return row[9]
                    elif datatype.upper().__contains__("Phone Number"):
                        return row[10]

        except FileNotFoundError:
            print("Database file not found.")
            return None

        print("User not found in the database.")
        return None

    def added_User_Data(self):
        employee_data = {}
        while True:
            print("\nMenu:")
            print("1. Added new Register Employee")
            print("2. Exit")

            choice = input("Enter your choice: ")
            if choice == "1":
                self.register_employee(employee_data)
            elif choice == "2":
                self.save_to_excel(employee_data)
                print("Exiting...")
                break
            else:
                print("Invalid choice. Try again.")

    def display_employee_data(self):
        emp = []
        max_col = self.ws.max_column
        for i in range(2, max_col + 1):
            cell_obj = self.ws.cell(row=1, column=i)
            emp.append(cell_obj.value)
            print(emp)

    def display_employee_data1(self):
        emp = []
        max_col = self.ws.max_column
        for i in range(1, self.ws.max_row + 1):
            print("Row ", i, " data :")

            for j in range(1, self.ws.max_column + 1):
                cell_obj = self.ws.cell(row=i, column=j)
                print(cell_obj.value, end="")
                print()

    # print(emp1)

    def display_records(self, employee_data):
        print("----------- User database records ------------------")
        for user_id, data in employee_data.items():
            print(f"1. User-ID:  {user_id}")
            print(f"2. Employee Id: {data['Employee Id']}")
            print(f"3. Name: {data['Name']}")
            print(f"4. Age: {data['Age']}")
            print(f"5. Company-Name: {data['Company-Name']}")
            print(f"6. Designation: {data['Designation']}")
            print(f"7. Salary: {data['Salary']}")
            print(f"8. Address: {data['Address']}")
            print(f"9. Phone Number: {data['Phone Number']}")
            print("")

    # def delete_rows_by_value(file_path, sheet_name, column_index, value_to_delete):
    #     # Load the Excel workbook
    #     workbook = load_workbook(file_path)
    #
    #     # Select the sheet
    #     sheet = workbook[sheet_name]
    #
    #     # Create a list to store the row indices to be deleted
    #     rows_to_delete = []
    #
    #     # Iterate through the rows and find the ones that contain the value to delete
    #     for row_index, row in enumerate(sheet.iter_rows(), start=1):
    #         cell_value = row[column_index - 1].value
    #         if cell_value == value_to_delete:
    #             rows_to_delete.append(row_index)
    #
    #     # Delete the identified rows in reverse order to avoid shifting issues
    #     for row_index in reversed(rows_to_delete):
    #         sheet.delete_rows(row_index)
    #
    #     # Save the modified Excel file
    #     workbook.save(file_path)
    #
    # # Example usage:
    # file_path = "path/to/your/excel/file.xlsx"
    # sheet_name = "Sheet1"
    # row_index = 2  # Assuming you want to delete rows based on values in the first column
    # value_to_delete = "ValueToDelete"
    # delete_rows_by_value(sheet_name, row_index, value_to_delete)
