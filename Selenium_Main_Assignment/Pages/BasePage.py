import time
import traceback

import openpyxl
from selenium.common import NoSuchElementException, ElementNotVisibleException, ElementNotSelectableException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from traceback import print_exc

from selenium.webdriver.common.by import By
from traceback import print_exc
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import *
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains


# create BasePage class
class BasePage:
    def __init__(self, driver):
        self.driver = driver

    def open_url(self, url):
        self.driver.get(url)

    # Custom wait for element visibility before interaction with it
    def do_wait(self, by_locator):
        WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(by_locator))

    # Get a list of web elements that match the given Xpath
    def get_list(self, xpath_string):
        by_locator = (By.XPATH, xpath_string)
        self.do_wait(by_locator)
        list_of_webelements = self.driver.find_elements(By.XPATH, xpath_string)
        return list_of_webelements

    # Move the mouse cursor to the given element
    def move_cursor(self, by_locator):
        actions = ActionChains(self.driver)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
        actions.move_to_element(self.driver.find_element(by_locator)).perform()

    # Move the mouse cursor to the required element identified by the given Xpath
    def move_cursor_to_required_element(self, xpath_string):
        by_locator = (By.XPATH, xpath_string)
        actions = ActionChains(self.driver)
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
        actions.move_to_element(self.driver.find_element(By.XPATH, xpath_string)).perform()

    # Custom wait for elements visibility before sending input text to it
    def do_send_keys(self, by_locator, input_text):
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator)).send_keys(input_text)

    # Get the visible text of the elements identified by the given locator
    def get_text(self, by_locator):
        element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
        return element.text

    # Check if an elements identified by the given locators is enabled and visible
    def is_enabled(self, by_locator):
        element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locator))
        return bool(element)

    # Get the page title and wait until it matches the provided title
    def get_page_title(self, title):
        WebDriverWait(self.driver, 10).until(EC.title_is(title))
        return self.driver.title

    def get_by_type(self, locator_type):

        locators = {'id': By.ID, 'name': By.NAME, 'xpath': By.XPATH,
                    'css': By.CSS_SELECTOR, 'class': By.CLASS_NAME, 'link': By.LINK_TEXT}
        if locator_type in locators:
            return locators.get(locator_type.lower())
        else:
            return False

    def wait_for_element(self, locator, timeout=10, poll_frequency=0.5):
        element = None
        try:

            wait = WebDriverWait(self.driver, 3, poll_frequency=1,
                                 ignored_exceptions=[NoSuchElementException,
                                                     ElementNotVisibleException,
                                                     ElementNotSelectableException])
            # wait.until(EC.visibility_of_element_located(locator))
            time.sleep(2)
        except:
            traceback.print_exc()
        return element

    # def waitForElement(self, locator):
    #     web_element = None
    #     try:
    #         wait = WebDriverWait(self.driver, 20, poll_frequency=1,
    #                              ignored_exceptions=[NoSuchElementException,
    #                                                  ElementNotVisibleException,
    #                                                  ElementNotSelectableException])
    #         web_element = wait.until(EC.visibility_of_element_located(locator))
    #     except Exception as e:
    #         traceback.print_exc()
    #     return web_element

    def get_element(self, locator, locator_type="id"):

        element = None
        try:
            locator_type = locator_type.lower()
            byType = self.get_by_type(locator_type)
            element = self.driver.find_element(byType, locator)
        except Exception as e:
            return element

    # def element_click(self, locator, locator_type="id"):
    #     element = None
    #     try:
    #         if locator:
    #             locator_type = locator_type.lower()
    #             self.wait_for_element(locator, locator_type)
    #             element = self.get_element(locator, locator_type)
    #         element.click()
    #     except Exception as e:
    #         print_exc()
    #         raise e

    def do_click(self, by_locater):
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locater)).click()

    def get_element_text(self, by_locater):
        element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(by_locater))
        return element.text

    def get_elements(self, locator, locator_type="id"):
        elements = None
        try:
            locator_type = locator_type.lower()
            byType = self.get_by_type(locator_type)
            elements = self.driver.find_elements(byType, locator)
        except Exception as e:
            print("exception failure")
        return elements

    def getTestData(self, testcaseName, filepath):
        data_dict = {}
        book = openpyxl.load_workbook(filepath)
        sheet = book.active

        # loop every row
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcaseName:
                # loop through each column
                for j in range(2, sheet.max_column + 1):
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [data_dict]
