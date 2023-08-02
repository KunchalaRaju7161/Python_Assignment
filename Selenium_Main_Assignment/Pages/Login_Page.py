import openpyxl
from selenium.webdriver.common.by import By

from Selenium_Main_Assignment.Configs.ConfigItems import ConfigItems
from Selenium_Main_Assignment.Pages.BasePage import BasePage


class Login_Page(BasePage):
    LOGIN = (By.XPATH, "//a[contains(text(),'Log in')]")
    USER_NAME = (By.XPATH, "//input[@id='loginusername']")
    PASSWORD = (By.XPATH, "//input[@id='loginpassword']")

    LOGIN_BUTTON = (By.XPATH, "//button[normalize-space()='Log in']")

    LOGIN_PROFILE = (By.XPATH, "//a[@id='nameofuser']")

    def __init__(self, driver):
        super().__init__(driver)
        self.driver = driver
        self.driver.maximize_window()

    def login_user_data(self, username, password):
        self.do_click(self.LOGIN)
        self.do_send_keys(self.USER_NAME, username)
        self.do_send_keys(self.PASSWORD, password)

    def login_user(self):
        bk = openpyxl.load_workbook(ConfigItems.Excel_FILE_PATH)
        sheet = bk["Sheet2"]

        username = sheet.cell(row=2, column=1).value
        password = sheet.cell(row=2, column=2).value

        print(username, password)
        self.login_user_data(username, password)
        self.do_click(self.LOGIN_BUTTON)
