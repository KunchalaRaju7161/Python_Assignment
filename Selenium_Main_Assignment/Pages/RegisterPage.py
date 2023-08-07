import openpyxl
from openpyxl import load_workbook
from selenium.webdriver.common.by import By

from Selenium_Main_Assignment.Pages.BasePage import BasePage
from Selenium_Main_Assignment.Configs.ConfigItems import ConfigItems


class RegisterPage(BasePage):
    PLACE_ORDER_BUTTON = (By.XPATH, "//button[contains(text(),'Place Order')]")
    NAME = (By.XPATH, "//input[@id='name']")
    COUNTRY = (By.XPATH, "//input[@id='country']")
    CITY = (By.XPATH, "//input[@id='city']")
    CARD = (By.XPATH, "//input[@id='card']")
    MONTH = (By.XPATH, "//input[@id='month']")
    YEAR = (By.XPATH, "//input[@id='year']")
    PURCHASE_BUTTON = (By.XPATH, "//button[contains(text(),'Purchase')]")

    THANK_YOU_MESSAGE = (By.XPATH, "//h2[contains(text(),'Thank you for your purchase!')]")
    PURCHASE_DETAILS_TEXT = (By.XPATH, "//p[@class='lead text-muted ']")
    OK_BUTTON = (By.XPATH, "//button[contains(text(),'OK')]")

    def __init__(self, driver):
        super().__init__(driver)
        self.driver = driver
        self.driver.maximize_window()

    def register_user_data(self, Name, Country, City, Credit_card, Month, Year):
        self.do_click(self.PLACE_ORDER_BUTTON)
        self.do_send_keys(self.NAME, Name)
        self.do_send_keys(self.COUNTRY, Country)
        self.do_send_keys(self.CITY, City)
        self.do_send_keys(self.CARD, Credit_card)
        self.do_send_keys(self.MONTH, Month)
        self.do_send_keys(self.YEAR, Year)
        self.do_click(self.PURCHASE_BUTTON)

    def register_user(self):
        bk = openpyxl.load_workbook(ConfigItems.EXCEL_FILE_PATH)
        sheet = bk["Sheet1"]

        Name = sheet.cell(row=2, column=1).value
        Country = sheet.cell(row=2, column=2).value
        City = sheet.cell(row=2, column=3).value
        Credit_card = sheet.cell(row=2, column=4).value
        Month = sheet.cell(row=2, column=5).value
        Year = sheet.cell(row=2, column=6).value
        print(Name, Country, City, Credit_card, Month, Year)
        self.register_user_data(Name, Country, City, Credit_card, Month, Year)

        thank_you_message_text = self.get_text(self.THANK_YOU_MESSAGE)
        purchase_detail_text = self.get_text(self.PURCHASE_DETAILS_TEXT)
        self.do_click(self.OK_BUTTON)
        return thank_you_message_text, purchase_detail_text
