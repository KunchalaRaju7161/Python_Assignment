import time

import openpyxl
from selenium.webdriver.common.by import By

from Selenium_Main_Assignment.Configs.ConfigItems import ConfigItems
from Selenium_Main_Assignment.Pages.BasePage import BasePage


class SignUp_Page(BasePage):
    SIGN_UP = (By.XPATH, "//a[contains(text(),'Sign up')]")
    USER_NAME = (By.XPATH, "//input[@id='sign-username']")
    PASSWORD = (By.XPATH, "//input[@id='sign-password']")

    SIGN_UP_BUTTON = (By.XPATH, "//button[normalize-space()='Sign up']")
    PROFILE = (By.XPATH, "//a[@id='nameofuser']")

    def __init__(self, driver):
        super().__init__(driver)
        self.driver = driver
        self.driver.maximize_window()

    def sign_up_user_data(self, username, password):
        self.do_click(self.SIGN_UP)
        self.do_send_keys(self.USER_NAME, username)
        self.do_send_keys(self.PASSWORD, password)

    def sign_up_user(self):
        bk = openpyxl.load_workbook(ConfigItems.Excel_FILE_PATH)
        sheet = bk["Sheet2"]

        username = sheet.cell(row=2, column=1).value
        password = sheet.cell(row=2, column=2).value

        print(username, password)
        self.sign_up_user_data(username, password)
        self.do_click(self.SIGN_UP_BUTTON)
        time.sleep(5)



